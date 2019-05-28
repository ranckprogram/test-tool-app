# -*- coding:utf-8 -*-
# -*- encoding:utf-8 -*-


import os
import sys
import openpyxl

reload(sys)
sys.setdefaultencoding("utf-8")


def read_case_id():
    """
    读取案例excel文件，获取案例基本信息：案例编号、案例名称、作者、创建日期
    :return: 以字典格式返回案例信息
    """
    v_dir = str(os.getcwd())  # 获取当前路径
    print v_dir
    dirs = os.listdir(v_dir)
    print dirs
    for file_name in dirs:  # 循环读取路径下的文件并筛选输出
        if file_name.endswith('.xlsm'):
            path = os.path.join(v_dir, file_name)
            print u"找到的案例文件为：", str(path)
        else:
            pass
    wb = openpyxl.load_workbook(path)  # 打开指定文件
    sheetnames = wb.get_sheet_names()
    print sheetnames
    ws = wb.get_sheet_by_name(sheetnames[2])
    # print "Work Sheet Titile:", ws.title
    # print "Work Sheet Rows:", ws.max_row
    # print "Work Sheet Cols:", ws.max_column

    # 建立存储数据的字典
    data_dic = {}
    for rx in range(4, ws.max_row + 1):
        case_id = ws.cell(row=rx, column=2).value  # 案例编号
        case_name = ws.cell(row=rx, column=3).value  # 案例名称
        print case_name
        case_creator = ws.cell(row=rx, column=12).value  # 案例作者
        case_time = ws.cell(row=rx, column=13).value  # 案例创建时间
        temp_list = [case_id, case_creator, case_time]
        if case_name:  # 将不为空的行写入字典数据
            data_dic[case_name] = temp_list
        else:
            break
    print data_dic

    # 打印字典数据个数
    print u"读取案例数:", len(data_dic)
    return data_dic


def create_rf_case(case_info):
    txt_name = "RF_Case.txt".encode('utf-8')  # 生成文件名字
    # if os.path.exists(txt_name):  # 首先判断文件是否存在，存在则删除。
    #     # os.remove(txt_name)
    #     pass
    # else:
    #     pass

    # 写入 Test Cases
    f = file(txt_name, "a+")
    a = f.readlines()
    keys = case_info.keys()
    keys.sort()  # 按照案例名称排序
    num = 0
    for key in keys:
        num = num + 1  # 统计写入案例数
        case_name = key + '\n'
        print case_name
        for index, value in enumerate(a):
            if value == case_name:
                case_id = u"     [Documentation]    *案例编号：* " + str(case_info[key][0]) + '\n    ...\n'  # 写入案例编号
                case_creator = u"    ...    *作者：* " + str(case_info[key][1]) + '\n' + "    ..." + '\n'  # 写入作者名称
                case_time = u"    ...    *日期：* " + str(case_info[key][2]) + '\n' + '\n'  # 写入创建时间
                a[index] = a[index] + case_id + case_creator + case_time
                print a[index]
    f.close()
    os.remove(txt_name)
    f = file(txt_name, "a+")
    f.writelines(a)
    f.close()
    print u"写入案例数:", num

if __name__ == '__main__':
    case_info = read_case_id()
    create_rf_case(case_info)
    raw_input("Press <enter>")
